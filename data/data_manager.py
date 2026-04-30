from __future__ import annotations
import json
import random
import openpyxl
from pathlib import Path
from typing import List, Optional, Tuple

from assets.constants import (
    SAVE_DIR, SAVE_FILE,
    PLAYER_XL_PATH, CPU_XL_PATH,
    POS_BATTERS, POS_ALL,
)
from game.player import Player, Roster, CPUTeam

PLAYER_XL: Path = PLAYER_XL_PATH
CPU_XL:    Path = CPU_XL_PATH


# ── 엑셀 읽기 ──────────────────────────────────────────
def _read_players(path: Path, sheet_name: str) -> List[Player]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h else "" for h in rows[0]]
    out = []

    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        d = dict(zip(header, r))
        name = d.get("이름")
        pos  = d.get("포지션")
        if not name or not pos:
            continue
        pos = str(pos).upper().strip()
        if pos not in POS_ALL:
            continue

        # 스탯은 3~7번째 열
        def get_stat(i):
            try:
                return int(r[i]) if i < len(r) and r[i] is not None else 60
            except (ValueError, TypeError):
                return 60

        try:
            p = Player(
                name=str(name).strip(),
                position=pos,
                power=get_stat(2),       # 파워 (타자) / 변화 (투수)
                accuracy=get_stat(3),    # 정확 (타자) / 구위 (투수)
                discipline=get_stat(4),  # 선구 (타자) / 제구 (투수)
                speed=get_stat(5),       # 주루 (타자) / 체력 (투수)
                defense=get_stat(6),     # 수비
                tier=str(d.get("티어", "B") or "B").upper(),
            )
        except Exception:
            continue
        out.append(p)
    return out


# ── 템플릿 생성 ────────────────────────────────────────
_PLAYER_HEADERS = ["이름", "포지션", "파워", "정확", "선구", "주루", "수비", "티어"]

def create_start_entry(path: Optional[Path] = None) -> Path:
    path = path or PLAYER_XL

    pool = _read_players(path, "선수풀")
    if not pool:
        raise ValueError("선수풀이 비어있음")

    wb = openpyxl.load_workbook(path)
    if "시작엔트리" in wb.sheetnames:
        del wb["시작엔트리"]
    ws = wb.create_sheet("시작엔트리")
    ws.append(_PLAYER_HEADERS)

    used = set()

    def pick(pos, tier):
        candidates = [p for p in pool if p.position == pos and p.tier == tier and p.pid not in used]
        if not candidates:
            candidates = [p for p in pool if p.position == pos and p.pid not in used]
        if not candidates:
            return None
        p = random.choice(candidates)
        used.add(p.pid)
        return p

    start_plan = [
        ("C",  "B"), ("1B", "B"), ("2B", "B"), ("3B", "B"),
        ("SS", "B"), ("LF", "B"), ("CF", "B"), ("RF", "B"), ("DH", "B"),
        ("C",  "C"), ("1B", "C"), ("2B", "C"), ("3B", "C"), ("LF", "C"),
        ("SP", "B"), ("SP", "B"), ("SP", "B"), ("SP", "C"), ("SP", "C"),
        ("RP", "B"), ("RP", "B"), ("RP", "C"), ("RP", "C"), ("RP", "C"), ("RP", "C"),
        ("CP", "B"),
    ]

    for pos, tier in start_plan:
        p = pick(pos, tier)
        if p:
            ws.append([p.name, p.position, p.power, p.accuracy,
                       p.discipline, p.speed, p.defense, p.tier])

    for col_idx in range(1, len(_PLAYER_HEADERS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    wb.save(path)
    return path


# 선수 풀 로드
def load_player_pool(path: Optional[Path] = None) -> List[Player]:
    return _read_players(path or PLAYER_XL, "선수풀")

# 시작 엔트리 로드
def load_start_roster(path: Optional[Path] = None) -> List[Player]:
    return _read_players(path or PLAYER_XL, "시작엔트리")

# cpu팀 로드
def load_cpu_teams(path: Optional[Path] = None) -> List[CPUTeam]:
    path = path or CPU_XL
    wb = openpyxl.load_workbook(path, data_only=True)

    if "팀목록" not in wb.sheetnames:
        return []
    ws = wb["팀목록"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]

    teams = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        d = dict(zip(header, r))
        team_name = str(d.get("팀명") or "").strip()
        if not team_name:
            continue
        try:
            diff  = int(d.get("난이도", 1) or 1)
            color = (int(d.get("R", 200) or 200),
                     int(d.get("G", 80)  or 80),
                     int(d.get("B", 80)  or 80))
        except (ValueError, TypeError):
            diff, color = 1, (200, 80, 80)

        roster = None
        if team_name in wb.sheetnames:
            players = _read_players(path, team_name)
            if players:
                roster = _assign_roster(players)
                roster.team_name = team_name

        teams.append(CPUTeam(name=team_name, difficulty=diff, color=color, roster=roster))
    return teams


# 로스터 자동 배정 
def _assign_roster(players: List[Player]) -> Roster:
    """선수 목록을 받아 포지션별로 자동 배정."""
    roster = Roster(team_name="My Team", gold=500)
    lineup_slots = {pos: None for pos in POS_BATTERS}
    bench, sp, rp, cp = [], [], [], []

    for p in players:
        if p.position in POS_BATTERS:
            if lineup_slots[p.position] is None:
                lineup_slots[p.position] = p
            elif len(bench) < 5:
                bench.append(p)
            else:
                roster.storage.append(p)
        elif p.position == "SP":
            if len(sp) < 5: sp.append(p)
            else: roster.storage.append(p)
        elif p.position == "RP":
            if len(rp) < 6: rp.append(p)
            else: roster.storage.append(p)
        elif p.position == "CP":
            if len(cp) < 1: cp.append(p)
            else: roster.storage.append(p)

    roster.lineup = [lineup_slots[pos] for pos in POS_BATTERS]
    roster.bench  = bench
    roster.sp, roster.rp, roster.cp = sp, rp, cp
    return roster


# 새 게임 / 세이브 
def new_game() -> Roster:
    create_start_entry()
    players = load_start_roster()
    return _assign_roster(players)


def save_game(roster: Roster, game_count: int = 0):
    SAVE_DIR.mkdir(parents=True, exist_ok=True)
    payload = {"roster": roster.to_dict(), "game_count": game_count}
    SAVE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_game() -> Optional[Tuple[Roster, int]]:
    if not SAVE_FILE.exists():
        return None
    try:
        data = json.loads(SAVE_FILE.read_text(encoding="utf-8"))
        if not data.get("roster"):
            return None
        return Roster.from_dict(data["roster"]), data.get("game_count", 0)
    except (json.JSONDecodeError, KeyError, TypeError):
        return None


# 뽑기 
def draw_from_pack(pool: List[Player], pack_type: str) -> Player:
    import copy
    from assets.constants import PACK_CHEAP_TIER_WEIGHTS, PACK_PREMIUM_TIER_WEIGHTS
    from game.player import new_id

    weights = PACK_CHEAP_TIER_WEIGHTS if pack_type == "cheap" else PACK_PREMIUM_TIER_WEIGHTS
    tiers = list(weights.keys())
    ws    = list(weights.values())
    chosen_tier = random.choices(tiers, weights=ws, k=1)[0]

    candidates = [p for p in pool if p.tier == chosen_tier] or pool
    drawn = copy.deepcopy(random.choice(candidates))
    drawn.pid = new_id()
    return drawn